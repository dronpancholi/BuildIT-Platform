"""
SEO Platform — Citation Export API
==================================
CSV and XLSX export for citation projects.
"""

from __future__ import annotations

import csv
import io
import uuid
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from seo_platform.core.auth import CurrentUser, get_current_user, get_validated_tenant_id
from seo_platform.core.database import get_session
from seo_platform.core.logging import get_logger
from seo_platform.models.citation_v2 import CitationProject, CitationSite, CitationSubmissionV2

logger = get_logger(__name__)
router = APIRouter()


async def _get_export_data(
    project_id: uuid.UUID,
    tenant_id: uuid.UUID,
    session: AsyncSession,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Fetch project + submissions data for export."""
    # Get project
    proj_result = await session.execute(
        select(CitationProject).where(
            CitationProject.id == project_id,
            CitationProject.tenant_id == tenant_id,
        )
    )
    project = proj_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get submissions with site info
    query = (
        select(CitationSubmissionV2, CitationSite)
        .join(CitationSite, CitationSubmissionV2.site_id == CitationSite.id)
        .where(
            CitationSubmissionV2.project_id == project_id,
            CitationSubmissionV2.tenant_id == tenant_id,
        )
        .order_by(CitationSite.difficulty_score.asc())
    )
    result = await session.execute(query)
    rows = result.all()

    project_data = {
        "business_name": project.business_name,
        "website_url": project.website_url or "",
        "category": project.category or "",
        "phone": project.phone or "",
        "email": project.email or "",
        "address": project.address or "",
        "city": project.city or "",
        "state": project.state or "",
        "country": project.country or "",
        "total_sites": project.total_sites,
    }

    submissions_data = []
    for sub, site in rows:
        submissions_data.append({
            "site_name": site.name,
            "site_url": site.url,
            "category": site.category,
            "difficulty": site.difficulty_score,
            "domain_authority": site.domain_authority,
            "status": sub.status,
            "listing_url": sub.listing_url or "",
            "account_created": "Yes" if sub.account_created else "No",
            "email_verified": "Yes" if sub.email_verified else "No",
            "listing_claimed": "Yes" if sub.listing_claimed else "No",
            "notes": sub.notes or "",
            "status_notes": sub.status_notes or "",
            "started_at": sub.started_at.isoformat() if sub.started_at else "",
            "submitted_at": sub.submitted_at.isoformat() if sub.submitted_at else "",
            "completed_at": sub.completed_at.isoformat() if sub.completed_at else "",
        })

    return project_data, submissions_data


# ---------------------------------------------------------------------------
# GET /citations/projects/{project_id}/export?format=csv
# ---------------------------------------------------------------------------
@router.get("/projects/{project_id}/export")
async def export_project(
    project_id: uuid.UUID,
    format: str = Query("csv", description="Export format: csv or xlsx"),
    tenant_id: uuid.UUID = Depends(get_validated_tenant_id),
    user: CurrentUser = Depends(get_current_user),
) -> StreamingResponse:
    """Export project submissions to CSV or XLSX."""
    async with get_session() as session:
        project_data, submissions_data = await _get_export_data(project_id, tenant_id, session)

        if format == "csv":
            return _export_csv(project_data, submissions_data)
        elif format == "xlsx":
            return _export_xlsx(project_data, submissions_data)
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'xlsx'.")


def _export_csv(
    project_data: dict[str, Any],
    submissions_data: list[dict[str, Any]],
) -> StreamingResponse:
    """Generate CSV export."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Site Name", "URL", "Category", "Status", "Listing URL",
        "Difficulty", "Domain Authority", "Notes", "Status Notes",
        "Account Created", "Email Verified", "Listing Claimed",
        "Started At", "Submitted At", "Completed At",
    ])

    # Rows
    for row in submissions_data:
        writer.writerow([
            row["site_name"],
            row["site_url"],
            row["category"],
            row["status"],
            row["listing_url"],
            row["difficulty"],
            row["domain_authority"],
            row["notes"],
            row["status_notes"],
            row["account_created"],
            row["email_verified"],
            row["listing_claimed"],
            row["started_at"],
            row["submitted_at"],
            row["completed_at"],
        ])

    output.seek(0)
    filename = f"{project_data['business_name'].replace(' ', '_')}_citations.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_xlsx(
    project_data: dict[str, Any],
    submissions_data: list[dict[str, Any]],
) -> StreamingResponse:
    """Generate XLSX export with openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="XLSX export requires openpyxl. Install with: pip install openpyxl",
        )

    wb = Workbook()

    # --- Sheet 1: Submissions ---
    ws = wb.active
    ws.title = "Submissions"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Site Name", "URL", "Category", "Status", "Listing URL",
        "Difficulty", "Domain Authority", "Notes", "Status Notes",
        "Account Created", "Email Verified", "Listing Claimed",
        "Started At", "Submitted At", "Completed At",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Status colors
    status_fills = {
        "new_backlink": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "already_exists": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "in_progress": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "failed": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "rejected": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }

    # Data rows
    for row_idx, row_data in enumerate(submissions_data, 2):
        values = [
            row_data["site_name"],
            row_data["site_url"],
            row_data["category"],
            row_data["status"],
            row_data["listing_url"],
            row_data["difficulty"],
            row_data["domain_authority"],
            row_data["notes"],
            row_data["status_notes"],
            row_data["account_created"],
            row_data["email_verified"],
            row_data["listing_claimed"],
            row_data["started_at"],
            row_data["submitted_at"],
            row_data["completed_at"],
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

        # Color the status cell
        status_cell = ws.cell(row=row_idx, column=4)
        fill = status_fills.get(row_data["status"])
        if fill:
            status_cell.fill = fill

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(submissions_data) + 2)
        )
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = min(max_length + 2, 40)

    # --- Sheet 2: Summary ---
    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ["Business Name", project_data["business_name"]],
        ["Website", project_data["website_url"]],
        ["Category", project_data["category"]],
        ["Phone", project_data["phone"]],
        ["Email", project_data["email"]],
        ["Address", project_data["address"]],
        ["City", project_data["city"]],
        ["State", project_data["state"]],
        ["Country", project_data["country"]],
        ["Total Sites", project_data["total_sites"]],
        ["", ""],
        ["Status", "Count"],
    ]

    # Count statuses
    status_counts: dict[str, int] = {}
    for row in submissions_data:
        s = row["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    for status, count in sorted(status_counts.items()):
        summary_data.append([status, count])

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 50

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{project_data['business_name'].replace(' ', '_')}_citations.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
